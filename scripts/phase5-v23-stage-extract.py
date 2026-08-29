#!/usr/bin/env python3
"""Strict read-only shared-folder staging and workbook structure extraction.

The source share is enumerated and hashed but never opened for write. Selected
workbooks are copied to project-local Git-ignored staging before parsing.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import sys
from pathlib import Path
from typing import Any

import openpyxl


BATCH_KEY = "phase5-v2.3-mx-history-001"
SUPPORTED = {".xlsx", ".xlsm"}
TF1_PATTERN = re.compile(r"(?:TENT\s+)?T?F1[-\s]*(?:PRE[-\s]*)?PEDIDO", re.I)


def utc_now() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def source_record(path: Path, share: Path, include_hash: bool = False) -> dict[str, Any]:
    stat = path.stat()
    record = {
        "relative_path": str(path.relative_to(share)),
        "filename": path.name,
        "size": stat.st_size,
        "last_modified_utc": dt.datetime.fromtimestamp(stat.st_mtime, dt.timezone.utc).isoformat(),
    }
    if include_hash:
        record["source_sha256"] = sha256(path)
    return record


def is_selected(relative_path: str, filename: str, suffix: str, po_prefix: str = "") -> bool:
    if suffix.lower() not in SUPPORTED or filename.startswith("~$"):
        return False
    normalized = relative_path.replace("/", "\\")
    if po_prefix and normalized.startswith(po_prefix.replace("/", "\\")):
        return True
    return TF1_PATTERN.search(filename) is not None


def inventory(share: Path, po_prefix: str = "") -> tuple[list[dict[str, Any]], list[Path]]:
    records: list[dict[str, Any]] = []
    selected: list[Path] = []
    for root, _, filenames in os.walk(share):
        for filename in filenames:
            if filename.startswith("~$") or filename in {"desktop.ini", "Thumbs.db"}:
                continue
            path = Path(root) / filename
            try:
                record = source_record(path, share)
            except OSError:
                continue
            records.append(record)
            if is_selected(record["relative_path"], filename, path.suffix, po_prefix):
                selected.append(path)
    records.sort(key=lambda item: item["relative_path"].casefold())
    selected.sort(key=lambda item: str(item).casefold())
    return records, selected


def stage(share: Path, selected: list[Path], staging: Path) -> list[dict[str, Any]]:
    files_dir = staging / "source-files"
    files_dir.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, Any]] = []
    for index, source in enumerate(selected, start=1):
        before = source_record(source, share, include_hash=True)
        local = files_dir / f"source-{index:03d}{source.suffix.lower()}"
        shutil.copy2(source, local)
        local_hash = sha256(local)
        after = source_record(source, share, include_hash=True)
        hash_verified = before["source_sha256"] == local_hash == after["source_sha256"]
        metadata_verified = before["size"] == after["size"] and before["last_modified_utc"] == after["last_modified_utc"]
        if not hash_verified or not metadata_verified:
            raise RuntimeError(f"Source changed during staging: {before['relative_path']}")
        manifest.append({
            "source_unc_path": str(source),
            "source_filename": source.name,
            "source_relative_path": before["relative_path"],
            "source_last_modified": before["last_modified_utc"],
            "source_size": before["size"],
            "source_sha256_before": before["source_sha256"],
            "local_staging_path": str(local),
            "local_sha256": local_hash,
            "source_sha256_after": after["source_sha256"],
            "hash_verified": hash_verified,
            "copied_at": utc_now(),
            "import_batch_id": BATCH_KEY,
        })
    return manifest


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value).replace("\u00a0", " ").strip()


def header_score(values: list[str], family: str) -> int:
    joined = " | ".join(value.upper() for value in values if value)
    if family == "CAVANNA_PO":
        tokens = ["SELLER", "BUYER", "PO", "STYLE", "DESCRIPTION", "QUANTITY", "QTY", "USD FOB", "DELIVERY DATE", "FABRIC"]
    else:
        tokens = ["SERIAL NUMBER", "PRODUCT NAME", "PRODUCT DETAILS", "ORDER VOLUME", "FACTORY PRICE", "CUSTOMER PRICE", "TOTAL CBM", "NUMBER OF BOXES"]
    return sum(token in joined for token in tokens)


def label_value(rows: list[list[Any]], label: str) -> Any:
    wanted = label.upper()
    for row in rows[:20]:
        values = [cell_text(value) for value in row]
        for index, value in enumerate(values):
            if value.upper() == wanted:
                return next((candidate for candidate in values[index + 1 :] if candidate), None)
    return None


def parse_workbook(entry: dict[str, Any]) -> dict[str, Any]:
    local = Path(entry["local_staging_path"])
    family = "TF1" if TF1_PATTERN.search(entry["source_filename"]) else "CAVANNA_PO"
    workbook = openpyxl.load_workbook(local, read_only=True, data_only=True, keep_links=False)
    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        scored = []
        for row_index, row in enumerate(rows[:80], start=1):
            values = [cell_text(value) for value in row]
            scored.append((header_score(values, family), row_index, values))
        score, header_row, headers = max(scored, default=(0, 0, []), key=lambda item: (item[0], -item[1]))
        if (family == "CAVANNA_PO" and score < 3) or (family == "TF1" and score < 3):
            sheets.append({
                "source_sheet": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "parse_status": "REVIEW",
                "reason": "NO_STABLE_HEADER",
            })
            continue
        data_rows = []
        for source_row, row in enumerate(rows[header_row:], start=header_row + 1):
            values = [cell_text(value) if not isinstance(value, (int, float)) else value for value in row]
            if not any(value not in (None, "") for value in values):
                continue
            data_rows.append({"source_row": source_row, "values": values})
        labels = {}
        if family == "CAVANNA_PO":
            labels = {
                "market": label_value(rows, "MEXICO") or ("MEXICO" if any(cell_text(value).upper() == "MEXICO" for row in rows[:10] for value in row) else None),
                "date": label_value(rows, "DATE:"),
                "client": label_value(rows, "CLIENT:"),
                "po": label_value(rows, "PO. NUMBER:"),
            }
        sheets.append({
            "source_sheet": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "parse_status": "PARSED",
            "header_row": header_row,
            "headers": headers,
            "labels": labels,
            "rows": data_rows,
        })
    workbook.close()
    return {
        "source_filename": entry["source_filename"],
        "source_relative_path": entry["source_relative_path"],
        "source_sha256": entry["local_sha256"],
        "source_last_modified": entry["source_last_modified"],
        "source_size": entry["source_size"],
        "family": family,
        "sheets": sheets,
    }


def compare_snapshots(before: list[dict[str, Any]], after: list[dict[str, Any]]) -> dict[str, Any]:
    old = {item["relative_path"]: item for item in before}
    new = {item["relative_path"]: item for item in after}
    deleted = sorted(set(old) - set(new))
    created = sorted(set(new) - set(old))
    modified = sorted(path for path in set(old) & set(new) if (old[path]["size"], old[path]["last_modified_utc"]) != (new[path]["size"], new[path]["last_modified_utc"]))
    return {
        "source_files_modified": len(modified),
        "source_files_deleted": len(deleted),
        "source_files_created": len(created),
        "source_files_renamed": 0,
        "source_files_moved": 0,
        "modified_paths": modified,
        "deleted_paths": deleted,
        "created_paths": created,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    configured_share = os.environ.get("DPV_SHARED_FOLDER_PATH")
    parser.add_argument("--share", default=configured_share, required=not configured_share)
    parser.add_argument("--po-prefix", default=os.environ.get("DPV_SHARED_PO_PREFIX", ""))
    parser.add_argument("--staging", required=True)
    args = parser.parse_args()
    share = Path(args.share)
    if not str(share).startswith("\\\\") or not share.is_dir():
        raise RuntimeError("The configured shared folder is not available or is outside the allowed source")
    if configured_share and str(share).casefold() != str(Path(configured_share)).casefold():
        raise RuntimeError("The requested shared folder does not match DPV_SHARED_FOLDER_PATH")
    staging = Path(args.staging).resolve()
    staging.mkdir(parents=True, exist_ok=True)
    before, selected = inventory(share, args.po_prefix)
    (staging / "source-snapshot-before.json").write_text(json.dumps(before, ensure_ascii=False, indent=2), encoding="utf-8")
    manifest = stage(share, selected, staging)
    (staging / "staging-manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    parsed = [parse_workbook(entry) for entry in manifest]
    (staging / "parsed-workbooks.json").write_text(json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8")
    after, _ = inventory(share, args.po_prefix)
    (staging / "source-snapshot-after.json").write_text(json.dumps(after, ensure_ascii=False, indent=2), encoding="utf-8")
    safety = compare_snapshots(before, after)
    (staging / "source-safety-result.json").write_text(json.dumps(safety, ensure_ascii=False, indent=2), encoding="utf-8")
    summary = {
        "batch_key": BATCH_KEY,
        "share_files_before": len(before),
        "share_files_after": len(after),
        "selected_files": len(selected),
        "hash_verified_files": sum(bool(item["hash_verified"]) for item in manifest),
        "parsed_files": len(parsed),
        "parsed_sheets": sum(sum(sheet["parse_status"] == "PARSED" for sheet in book["sheets"]) for book in parsed),
        "review_sheets": sum(sum(sheet["parse_status"] == "REVIEW" for sheet in book["sheets"]) for book in parsed),
        "safety": safety,
        "completed_at": utc_now(),
    }
    (staging / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
